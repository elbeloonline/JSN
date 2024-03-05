

from enum import Enum
import logging
import os

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import connection

from openpyxl import load_workbook, Workbook

class MissingDataSheets:
    JudgmentDetails = 0
    PartySummary = 1
    DebtSummary = 2
    JudgmentSummary = 3
    DocumentSummary = 4

class Command(BaseCommand):
    help = 'Load missing judgments into system that were scraped by Transdata'

    def __init__(self):
        self.logger = logging.getLogger(__name__)

    def load_xlsx_file(self, xlsx_filename):
        """
        Load a xlsx file and return a handle to it as a workbook
        :param xlsx_filename:
        :type xlsx_filename: str
        :return: openpyxl.Workbook
        """
        xlsx_pathname = os.path.join(settings.MEDIA_ROOT, xlsx_filename)
        wb = load_workbook(xlsx_pathname)
        return wb

    def _select_workbook_sheet(self, wb, sheetname):
        from openpyxl.worksheet import worksheet
        # type: (Workbook, int) -> worksheet
        """
        Select a sheet from the provided workbook
        :param wb: the workbook to work with
        :param sheetname: the sheetname to select
        :return: the worksheet corresponding to the provided name
        """
        ws = wb.worksheets[sheetname]
        return ws

    def _load_data_from_worksheet(self, ws):
        from openpyxl.worksheet import worksheet
        # type: (worksheet) -> List[str]
        """
        extract data from worksheet into a list
        :param ws: the worksheet to parse
        :return: List of judgment rows. These can be converted one by one into the corresponding database object
        """
        # process header
        for ws_row in ws.iter_rows(min_row=1, min_col=1, max_row=1):
            written_judgment_row_header = []
            for ws_cell in ws_row:
                written_judgment_row_header.append(ws_cell.value)
        num_cols = len(written_judgment_row_header)
        # process remaining data
        written_judgment_data = []
        for ws_row in ws.iter_rows(min_row=2, min_col=1, max_col=num_cols):
            written_judgment_row_data = []
            for ws_cell in ws_row:
                written_judgment_row_data.append(ws_cell.value)
            if not written_judgment_row_data[0] == None:
                written_judgment_data.append(written_judgment_row_data)
        return written_judgment_data

    def load_judgment_details(self, wb, ws):
        # type: (Workbook) -> None
        ws_data = self._load_data_from_worksheet(ws)
        return ws_data

    def _data_to_judgmentdetails_list(self, j_row_data_list):
        from statereport.models import MissingJudgmentDetails
        j_obj_list = []
        for rd in j_row_data_list:
            o = MissingJudgmentDetails(JudgmentSummary=rd[0], Name=rd[1], RelatedJudgment=rd[2],
                                       DocketNumber=rd[3], VenueId=rd[4], FilingLocation=rd[5], Court=rd[6], JudgmentStatus=rd[7],
                                       StatusDate=rd[8], JudgmentAmount=rd[9], CourtCosts=rd[10], Interest=rd[11],
                                       AttorneyFee=rd[12], OtherAmount=rd[13], ProcessingLocation=rd[14],
                                       JudgmentEnterDate=rd[15], Time=rd[16], JudgmentFilingDate=rd[17])
            o.JudgmentNum = self._judgment_num_from_summary_string(rd[0])
            j_obj_list.append(o)
        return j_obj_list

    def _data_to_partysummary_list(self, j_row_data_list):
        from statereport.models import MissingJudgmentParty
        j_obj_list = []
        for rd in j_row_data_list:
            o = MissingJudgmentParty(JudgmentSummary=rd[0], DebtID=rd[1], DebtStatus=rd[2], DebtAmount=rd[3],
                                     AttorneyFee=rd[4], Cost=rd[5], Interest=rd[6], OtherAmount=rd[7], DebtEnterDate=rd[8])
            o.JudgmentNum = self._judgment_num_from_summary_string(rd[0])
            j_obj_list.append(o)
        return j_obj_list

    def _data_to_debtsummary_list(self, j_row_data_list):
        from statereport.models import MissingJudgmentDebtSummary
        j_obj_list = []
        for rd in j_row_data_list:
            o = MissingJudgmentDebtSummary(JudgmentSummary=rd[0], DebtID=rd[1], Name=rd[2], Role=rd[3],
                                           AlternateNames=rd[4], DebtPartyStatus=rd[5], StatusDate=rd[6])
            o.JudgmentNum = self._judgment_num_from_summary_string(rd[0])
            j_obj_list.append(o)
        return j_obj_list

    def _data_to_judgmentsummary_list(self, j_row_data_list):
        from statereport.models import MissingJudgmentJudgmentSummary
        j_obj_list = []
        for rd in j_row_data_list:
            o = MissingJudgmentJudgmentSummary(JudgmentSummary=rd[0], DebtID=rd[1], JudgmentAmount=rd[2], JudgmentStatus=rd[3],
                                               StatusDate=rd[4], DebtAmount=rd[5], DebtStatus=rd[6], PartyDebtStatus=rd[7],
                                               PartyDebtStatusDate=rd[8], PartyInformation=rd[9], Role=rd[10], AlternateNames=rd[11],
                                               SelfRepresented=rd[12], BirthDate=rd[13], Address1=rd[14], Address2=rd[15], Phone=rd[16],
                                               Attorney=rd[17], Name=rd[18], GuardianInformation=rd[19], GuardianAddress1=rd[20],
                                               GuardianAddress2=rd[21], AffiliationCode=rd[22], AppointmentDate=rd[23],
                                               ReqBondAmount=rd[24], BondReceivedIndicator=rd[25])
            o.JudgmentNum = self._judgment_num_from_summary_string(rd[0])
            j_obj_list.append(o)
        return j_obj_list

    def _data_to_documentsummary_list(self, j_row_data_list):
        from statereport.models import MissingJudgmentDocumentSummary
        j_obj_list = []
        for rd in j_row_data_list:
            o = MissingJudgmentDocumentSummary(JudgmentSummary=rd[0], DebtId=rd[1], DocumentType=rd[2], DocumentFileDate=rd[3],
                                               DocumentStatus=rd[4], PartyNameRole1=rd[5], PartyNameRole2=rd[6],
                                               PartyNameRole3=rd[7], PartyNameRole4=rd[8])
            o.JudgmentNum = self._judgment_num_from_summary_string(rd[0])
            j_obj_list.append(o)
        return j_obj_list

    def save_objects_to_db_from_sheet_data(self, sheet_type, ws_data):
        from statereport.models import MissingJudgmentDetails, MissingJudgmentParty, MissingJudgmentDebtSummary, MissingJudgmentJudgmentSummary, MissingJudgmentDocumentSummary
        if sheet_type == MissingDataSheets.JudgmentDetails:
            obj_list = self._data_to_judgmentdetails_list(ws_data)
            MissingJudgmentDetails.objects.bulk_create(obj_list)
        elif sheet_type == MissingDataSheets.PartySummary:
            obj_list = self._data_to_partysummary_list(ws_data)
            MissingJudgmentParty.objects.bulk_create(obj_list)
        elif sheet_type == MissingDataSheets.DebtSummary:
            obj_list = self._data_to_debtsummary_list(ws_data)
            MissingJudgmentDebtSummary.objects.bulk_create(obj_list)
        elif sheet_type == MissingDataSheets.JudgmentSummary:
            obj_list = self._data_to_judgmentsummary_list(ws_data)
            MissingJudgmentJudgmentSummary.objects.bulk_create(obj_list)
        elif sheet_type == MissingDataSheets.DocumentSummary:
            obj_list = self._data_to_documentsummary_list(ws_data)
            MissingJudgmentDocumentSummary.objects.bulk_create(obj_list)

    def _judgment_num_from_summary_string(self, str):
        """
        Extract the judgment number from a descriptive string. Assumes the judgment number is at the beginning of the string
        :param str:
        :return:
        """
        s = str.split('-')[0].strip()
        s = s.replace(' ', '-')
        return s

    def handle(self, *args, **options):
        wb = self.load_xlsx_file('Open Judgements Truncated.xlsx')
        ws = self._select_workbook_sheet(wb, MissingDataSheets.JudgmentDetails)
        ws_data = self.load_judgment_details(wb, ws)
        self.save_objects_to_db_from_sheet_data(MissingDataSheets.JudgmentDetails, ws_data)

        ws = self._select_workbook_sheet(wb, MissingDataSheets.PartySummary)
        ws_data = self.load_judgment_details(wb, ws)
        self.save_objects_to_db_from_sheet_data(MissingDataSheets.PartySummary, ws_data)

        ws = self._select_workbook_sheet(wb, MissingDataSheets.DebtSummary)
        ws_data = self.load_judgment_details(wb, ws)
        self.save_objects_to_db_from_sheet_data(MissingDataSheets.DebtSummary, ws_data)

        ws = self._select_workbook_sheet(wb, MissingDataSheets.JudgmentSummary)
        ws_data = self.load_judgment_details(wb, ws)
        self.save_objects_to_db_from_sheet_data(MissingDataSheets.JudgmentSummary, ws_data)

        # ws = self._select_workbook_sheet(wb, MissingDataSheets.DocumentSummary)
        # ws_data = self.load_judgment_details(wb, ws)
        # self.save_objects_to_db_from_sheet_data(MissingDataSheets.DocumentSummary, ws_data)

        self.logger.info(ws_data)
        self.logger.info("Done")

