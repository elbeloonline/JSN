import os

from django.conf import settings

from nameviewer.printfactory import PrintFactory
from .utils import SpecCodes


class JSNCSVWriter:

    @staticmethod
    def delete_csv(order_id):
        """
        Remove a previously generated CSV from the file system that's associated with an order id
        :param order_id: order id of the file to be removed
        """
        import os
        from orders.models import Order

        # client_number = Order.objects.get(id=order_id).title_number
        client_number = '{}-{}'.format(Order.objects.get(id=order_id).title_number, order_id)
        csv_filename = "{}.csv".format(client_number)
        xlsx_filename = "{}.xlsx".format(client_number)
        csv_file = os.path.join(os.path.join(settings.MEDIA_ROOT, csv_filename))
        if os.path.exists(csv_file):
            os.remove(csv_file)
        xlsx_file = os.path.join(os.path.join(settings.MEDIA_ROOT, xlsx_filename))
        if os.path.exists(xlsx_file):
            os.remove(xlsx_file)

    @staticmethod
    def extract_creditor_party(case_instance, pf):
        """
        Get a creditor's party name from a case instance
        :param case_instance: Case instance to use for obtaining party data
        :param pf: PrintFactory instance to use for formatting name
        :return: str of formatted party name
        """
        party_name = None
        for party_instance in case_instance.party_set.all():
            if party_instance.party_role_type_code == "C":
                party_name = pf.format_full_party_name(party_instance)
                break
        return party_name

    @staticmethod
    def get_csv_header():
        """
        Generates the CSV header needed for writing search results to a csv file
        :return: header contents
        :type list
        """
        # csv_header = ['Matched Names', 'Also Known As', 'Document Number', 'Page Number', 'Case Number', 'County', 'DOB',
        #               'Last 4 Digits of SS', 'Partial DL', 'Date Entered', 'Type', 'Creditor', 'Amount', 'Lawyer', 'Lawyer Phone',
        #               'BK Discharge Date', 'BK Final Decree Date', 'Omit']
        csv_header = ['Document Number', 'Matched Names', 'DOB', 'Last 4 Digits of SS', 'Partial DL', 'Address', 'Filing Location', 'Case Title', 'Filing Date', 'Amount', 'Page Number']
        return csv_header

    @staticmethod
    def write_scnj_csv(order_id, cases, search_name, csv_filename, judgment_dobs, is_high_value=False):
        """
        Generate a csv/xlsx file for an order
        :param order_id: Order id to use for generating the file
        :param cases: A list of cases associated with the order id
        :param search_name: the original name searched
        :param csv_filename: the name of the csv file to write to
        :param judgment_dobs: a list of DOBs associated with judgments belonging to a list of cases for the order id
        :return:
        """
        from .helpers import first_name_variations_from_db, last_name_variations_from_db
        from nameviewer.pdfutils import JudgmentPageExtractor, PdfMerger
        from statereport.utils import StateReportNameSearchUtils

        JSNCSVWriter.check_reset_csv_file_data(order_id)

        search_fname = search_name.first_name
        search_lname = search_name.last_name

        fname_vars = first_name_variations_from_db(search_fname)
        fname_vars = [x.lower() for x in fname_vars]
        lname_vars = last_name_variations_from_db(search_lname)

        judgments_list = JSNCSVWriter.judgments_from_cases(cases)  # -> List
        state_pdf_filename = PdfMerger.get_state_report_from_order_id(order_id, is_high_value)
        judgment_page_dict = JudgmentPageExtractor.extractScnjPages(judgments_list, state_pdf_filename)
        judgments_page_num_offset = JudgmentPageExtractor.countIndexPages(order_id)
        # judgments_page_num_offset = 0

        print("=-=-=-=-=-= SCNJ CSV Writer Results =-=-=-=-=-=")
        pf = PrintFactory()
        csv_lines = []
        for case_instance in cases:
            docket_number = JSNCSVWriter._make_docket_number_from_case(case_instance)
            case_number = case_instance.acms_docket_number
            action_type = pf._map_civil_venue_code(case_instance.case_type_code, SpecCodes.CIVIL)
            county = case_instance.acms_venue_id
            creditor_party = JSNCSVWriter.extract_creditor_party(case_instance, pf)
            for debt_instance in case_instance.debt_set.all():
                debt_amt = pf._make_float(debt_instance.party_orig_amt, False)
                # print('Amount: ' + debt_amt)
                actual_case_date = case_instance.case_filed_date
                if actual_case_date == '00000000':  # @TODO: doesn't work well for multiple debt instances
                    actual_case_date = debt_instance.entered_date
                case_filed_date = pf._make_american_date(actual_case_date)

            scnj_page_num = judgment_page_dict[docket_number]
            case_title = case_instance.case_title
            parties_listed = []
            all_parties = []
            alt_party_to_add = None
            atty_name = ""
            addr = ""
            party_to_add = None
            for party_instance in case_instance.party_set.all():
                party_name = party_instance.party_last_name + ' ' + party_instance.party_first_name + ' ' + party_instance.party_initial
                all_parties.append(party_name)
                if set(fname_vars).intersection(set(party_name.lower().split())) and set(lname_vars).intersection(set(party_name.lower().split())):
                    party_name_to_add = party_name
                    party_to_add = party_instance
                    atty_name = party_instance.atty_firm_first_name.strip() + " " + party_instance.atty_firm_last_name.strip()
                    break
            if not party_to_add:
                parties_listed = all_parties
            else:
                parties_listed.append(party_name_to_add.strip())
                addr = PrintFactory.make_party_address(party_to_add)

            all_alt_parties = []
            alt_parties_listed = []
            for alt_party_instance in case_instance.partyalt_set.all():
                alt_party_name = alt_party_instance.party_last_name + ' ' + alt_party_instance.party_first_name + ' ' + alt_party_instance.party_initial
                all_alt_parties.append(alt_party_name)
                if set(fname_vars).intersection(set(alt_party_name.lower().split())) and set(lname_vars).intersection(set(alt_party_name.lower().split())):
                    alt_party_to_add = alt_party_name.strip()
                    # atty_name = party_instance.atty_firm_first_name.strip() + " " + party_instance.atty_firm_last_name.strip()
                    break
            if not alt_party_to_add:
                alt_parties_listed.append("")
            else:
                alt_parties_listed.append(alt_party_to_add)
            # judgment_dob = judgment_dobs.get(docket_number, '')
            nd_docket_number = docket_number.replace('-','')
            dob_key = nd_docket_number[:-4] + nd_docket_number[-2:]
            judgment_dob = judgment_dobs.get(dob_key, '')

            dlicense = ""
            partial_ssn = ""
            if settings.MERGE_PDF_SCRAPED_PARTY_INFO:
                scraped_judg = StateReportNameSearchUtils.get_scraped_judgment_data(docket_number)
                if scraped_judg:
                    if scraped_judg.party_dlicense:
                        dlicense = scraped_judg.party_dlicense
                    if scraped_judg.party_ssn:
                        partial_ssn = scraped_judg.party_ssn

            # print('Party names:' + "; ".join(parties_listed))
            csv_item = []
            csv_parties = "; ".join(parties_listed)
            csv_alt_parties = "; ".join(alt_parties_listed)
            # csv_item.append(csv_alt_parties)  # Also known as
            csv_item.append(docket_number)  # Docket Number
            csv_item.append(csv_parties)
            csv_item.append(judgment_dob)  # DOB
            csv_item.append(partial_ssn)  # Last 4 SS#/ITIN
            csv_item.append(dlicense)  # Partial DL #
            csv_item.append(addr)  # Address
            csv_item.append(county)  # County/Filing Location
            # csv_item.append('')  # Party Role
            csv_item.append(case_title)  # Case Title
            csv_item.append(case_filed_date)  # Date Entered
            csv_item.append(debt_amt)
            csv_item.append(int(scnj_page_num) + judgments_page_num_offset)  # Docket Number Page Num
            # csv_item.append(case_number)  # Case Number
            # csv_item.append(action_type)  # judgment type
            # csv_item.append(creditor_party)  # Creditor
            # csv_item.append(debt_amt)
            # csv_item.append(atty_name)  # Lawyer
            # csv_item.append('')  # Lawyer Phone Number
            # csv_item.append('')  # Discharge
            # csv_item.append('')  # Final Decree
            # csv_item.append('')  # Omit
            csv_lines.append(csv_item)
        # print(csv_lines)

        # import csv
        import xlsxwriter
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import PatternFill

        xlsx_filename = csv_filename.replace('.csv', '.xlsx')
        xlsx_pathname = os.path.join(settings.MEDIA_ROOT, xlsx_filename)

        try:
            wb = load_workbook(xlsx_pathname)
            ws = wb.worksheets[0]
        except IOError:  # only add header to worksheet when creating a new file
            csv_header = JSNCSVWriter.get_csv_header()
            wb = Workbook()
            ws = wb.active
            ws.append(csv_header)

        # extract data from worksheet to avoid writing duplicate judgments
        written_judgments = []
        for ws_row in ws.iter_rows(min_row=1, min_col=1, max_col=1):
            for ws_cell in ws_row:
                written_judgments.append(ws_cell.value)

        # write data
        fill_color = "DAEEF3"
        for line_num, line_data in enumerate(csv_lines):
            if not line_data[0] in written_judgments:  # don't write dupes
                if line_num % 2 == 1:
                    # color line
                    for cell in ws[line_num+1]:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                ws.append(line_data)

        # adjust column widths
        from openpyxl.styles import Alignment
        max_col_width = 30
        for column_cells in ws.columns:
            # length = max(len(str(cell.value)) for cell in column_cells)
            length = 0
            for cell in column_cells:
                tmp_length = len(str(cell.value))
                if tmp_length > length:
                    length = tmp_length
                cell.alignment= Alignment(wrap_text=True)
            if length > max_col_width:
                length = max_col_width
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2


        wb.save(xlsx_pathname)
        wb.close()

        print(("Wrote file to {}".format(xlsx_filename)))
        print("=-=-=-=-=-= CSV Writer Results =-=-=-=-=-=")

    @staticmethod
    def write_bk_csv(top_xml_element, search_name, order, matched_cases_dict):
        """
        Generate a CSV that can be downloaded by the end user to see matched BKCY cases
        :param top_xml_element: list of matched bankruptcy cases
        :type top_xml_element: xml.etree.ElementTree.Element
        :param search_name: search name supplied by end user
        :type search_name: orders.model.Searchname
        :param order: original order object
        :type order: orders.Order
        :return: None
        """
        from openpyxl.styles import PatternFill
        from .helpers import first_name_variations_from_db, last_name_variations_from_db
        from nameviewer.pdfutils import JudgmentPageExtractor, PdfMerger

        search_fname = search_name.first_name
        search_lname = search_name.last_name

        fname_vars = first_name_variations_from_db(search_fname)
        fname_vars = [x.lower() for x in fname_vars]
        lname_vars = last_name_variations_from_db(search_lname)

        judgments_list = JSNCSVWriter.bk_judgment_nums_from_cases(top_xml_element)  # -> List
        bkcy_pdf_filename = PdfMerger.get_bkcy_report_from_order_id(order.id)
        # xlsx_filename = order.title_number + ".xlsx"
        xlsx_filename = "{}-{}.xlsx".format(order.title_number, order.id)
        judgment_page_dict = JudgmentPageExtractor.extractBkcyPages(judgments_list, bkcy_pdf_filename, xlsx_filename)

        print("=-=-=-=-=-= BK CSV Writer Results =-=-=-=-=-=")
        pf = PrintFactory()
        csv_lines = []
        for case_instance in top_xml_element.findall('case'):
            try:
                judgment_type_code = 'Bankruptcy'
                docket_number = case_instance.find('BANKRUPTCY_NUMBER').text  # BKCY number
                bkcy_page_num = judgment_page_dict[docket_number]
                action_type = case_instance.find('CHAPTER').text    # BKCY chapter
                case_filed_date = case_instance.find('DATE_FILED').text
                parties_element = case_instance.find('PARTIES')
                atty_text_list = parties_element.find('ATTORNEY').text.split('\n')
                atty_name = atty_text_list[0]
                case_title = matched_cases_dict[docket_number]['case_model'].case_title

                import re
                atty_phone = ''
                for atty_line in atty_text_list:
                    m = re.search('\d+\-\d+', atty_line)
                    if m and len(m.group(0)) > 0:
                        atty_phone = atty_line

                bkcy_party = parties_element.find('DEBTOR').text
                case_discharge_date = case_instance.find('DATE_DISCHARGED').text
                case_final_decree_date = case_instance.find('DATE_TERMINATED').text

                csv_item = []
                # csv_item.append('')  # Also known as
                csv_item.append(docket_number)  # BKCY number
                csv_item.append(bkcy_party)
                csv_item.append('')  # DOB
                csv_item.append('')  # Address
                csv_item.append('')  # County/Filing Location
                # csv_item.append('')  # Party Role
                csv_item.append('')  # Last 4 SS#/ITIN
                csv_item.append('')  # Partial DL #
                csv_item.append(case_title)  # Case Title
                csv_item.append(case_filed_date)  # Date Entered
                csv_item.append('')  # debt amount
                csv_item.append(int(bkcy_page_num))  # page number
                # csv_item.append('')  # no case number
                # csv_item.append(judgment_type_code)  # @TODO: make Bankruptcy/Voluntary
                # csv_item.append('')  # Creditor
                # csv_item.append('')
                # csv_item.append(atty_name)  # Lawyer
                # csv_item.append(atty_phone)  # Lawyer Phone Number
                # csv_item.append(case_discharge_date)  # Discharge
                # csv_item.append(case_final_decree_date)  # Final Decree
                # csv_item.append('')  # Omit
                csv_lines.append(csv_item)
            except:  # Usually happens if the BKCY couldn't be located
                docket_number = case_instance.find('BANKRUPTCY_NUMBER').text
                print(("Error while trying to write csv for BKCY {}".format(docket_number)))
                # partial_item = [''] * 2 + [docket_number] + [''] * 15
                partial_item = [''] * 2 + [docket_number] + [''] * 7 + [judgment_type_code] + [''] * 7
                # csv_lines.append(partial_item)
        print(csv_lines)


        from openpyxl import load_workbook, Workbook

        # xlsx_filename = "{}-{}.xlsx".format(order.title_number, order.id)
        xlsx_pathname = os.path.join(settings.MEDIA_ROOT, xlsx_filename)

        try:
            wb = load_workbook(xlsx_pathname)
            ws = wb.worksheets[0]
        except (OSError, IOError) as e:
            csv_header = JSNCSVWriter.get_csv_header()
            wb = Workbook()
            ws = wb.active
            ws.append(csv_header)

        # write data
        fill_color = "DAEEF3"
        for line_num, line_data in enumerate(csv_lines):
            if line_num % 2 == 1:
                # color line
                for cell in ws[line_num+1]:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                ws.append(line_data)
        wb.save(xlsx_pathname)

        print(("Wrote file to {}".format(xlsx_filename)))
        print("=-=-=-=-=-= CSV Writer Results =-=-=-=-=-=")

    @staticmethod
    def write_usdc_csv(top_xml_element, search_name, order, matched_cases_dict):
        """
        Generate a CSV that can be downloaded by the end user to see matched BKCY cases
        :param top_xml_element: list of matched bankruptcy cases
        :type top_xml_element: xml.etree.ElementTree.Element
        :param search_name: search name supplied by end user
        :type search_name: orders.model.Searchname
        :param order: original order object
        :type order: orders.Order
        :return: None
        """
        from .helpers import first_name_variations_from_db, last_name_variations_from_db
        from nameviewer.pdfutils import JudgmentPageExtractor, PdfMerger

        search_fname = search_name.first_name
        search_lname = search_name.last_name

        fname_vars = first_name_variations_from_db(search_fname)
        fname_vars = [x.lower() for x in fname_vars]
        lname_vars = last_name_variations_from_db(search_lname)

        judgments_list = JSNCSVWriter.usdc_judgment_nums_from_cases(top_xml_element)  # -> List
        usdc_pdf_filename = PdfMerger.get_usdc_report_from_order_id(order.id)
        xlsx_filename = "{}-{}.xlsx".format(order.title_number, order.id)
        judgment_page_dict = JudgmentPageExtractor.extractUsdcPages(judgments_list, usdc_pdf_filename, xlsx_filename)

        print("=-=-=-=-=-= USDC CSV Writer Results =-=-=-=-=-=")
        pf = PrintFactory()
        csv_lines = []
        for case_instance in top_xml_element.findall('case'):
            try:
                judgment_type_code = 'USDC'
                docket_number = case_instance.find('USDC_CASE_NUMBER').text  # USDC number
                usdc_page_num = judgment_page_dict[docket_number]
                # action_type = case_instance.find('CHAPTER').text    # BKCY chapter
                case_filed_date = case_instance.find('DATE_FILED').text
                parties_element = case_instance.find('PARTIES')
                atty_text_list = parties_element.find('ATTORNEY').text.split('\n')
                atty_name = atty_text_list[0]
                # case_title = matched_cases_dict[docket_number]['case_model'].case_title

                import re
                atty_phone = ''
                for atty_line in atty_text_list:
                    m = re.search('\d+\-\d+', atty_line)
                    if m and len(m.group(0)) > 0:
                        atty_phone = atty_line

                usdc_parties = parties_element.find('DEBTOR').text.split('\n')
                # parse the right party name from the multiple lines of text
                usdc_party = ''
                if len(usdc_parties) > 0:
                    usdc_party = usdc_parties[0]
                for party in usdc_parties:
                    party_elements = party.split(' ')
                    for party_name_element in party_elements:
                        if party_name_element.lower() in fname_vars:
                            usdc_party = party
                case_discharge_date = case_instance.find('DATE_DISCHARGED').text
                case_final_decree_date = case_instance.find('DATE_TERMINATED').text

                csv_item = []
                # csv_item.append('')  # Also known as
                csv_item.append(docket_number)  # USDC number
                csv_item.append(usdc_party)  # @TODO: fixme
                csv_item.append('')  # DOB
                csv_item.append('')  # Address
                csv_item.append('')  # County

                # csv_item.append('')  # Party Role
                csv_item.append('')  # Case Title - @TODO: need to convert to native format in original scraped data
                csv_item.append(case_filed_date)  # Date Entered
                csv_item.append('')  # debt amount
                csv_item.append(usdc_page_num)  # page number
                # csv_item.append('')  # no case number
                csv_item.append('')  # Last 4 SS#/ITIN
                csv_item.append('')  # Partial DL #
                # csv_item.append(judgment_type_code)
                # csv_item.append('')  # Creditor
                # csv_item.append('')
                # csv_item.append(atty_name)  # Lawyer
                # csv_item.append(atty_phone)  # Lawyer Phone Number
                # csv_item.append(case_discharge_date)  # Discharge
                # csv_item.append(case_final_decree_date)  # Final Decree
                # csv_item.append('')  # Omit
                csv_lines.append(csv_item)
            except:  # Usually happens if the BKCY couldn't be located
                docket_number = case_instance.find('USDC_CASE_NUMBER').text
                print(("Error while trying to write csv for USDC {}".format(docket_number)))
                # partial_item = [''] * 2 + [docket_number] + [''] * 15
                partial_item = [''] * 2 + [docket_number] + [''] * 7 + [judgment_type_code] + [''] * 7
                # csv_lines.append(partial_item)
        print(csv_lines)

        from openpyxl import load_workbook, Workbook

        # xlsx_filename = "{}.xlsx".format(order.title_number)
        xlsx_pathname = os.path.join(settings.MEDIA_ROOT, xlsx_filename)

        try:
            wb = load_workbook(xlsx_pathname)
            ws = wb.worksheets[0]
        except OSError:
            csv_header = JSNCSVWriter.get_csv_header()
            wb = Workbook()
            ws = wb.active
            ws.append(csv_header)

        for line in csv_lines:
            ws.append(line)
        wb.save(xlsx_pathname)

        print(("Wrote file to {}".format(xlsx_filename)))
        print("=-=-=-=-=-= CSV Writer Results =-=-=-=-=-=")

    @staticmethod
    def judgments_from_cases(cases):
        """
        Generate a list of docketed judgment numbers from a list of cases
        :param cases: list of cases to iterate through when generating judgment numbers
        :return: list of judgment numbers
        """
        judgment_list = []
        for case_instance in cases:
            docketed_judgment = JSNCSVWriter._make_docket_number_from_case(case_instance)
            judgment_list.append(docketed_judgment)
        return judgment_list

    @staticmethod
    def _make_docket_number_from_case(case_instance):
        """
        Generate a docket number from a case
        :param case_instance: the case to use for extracting the docket number
        :return: a docket number in the format of judgment code-seq num-YYYY
        """
        judgment_type_code = case_instance.docketed_judgment_type_code
        seq_num = case_instance.docketed_judgment_seq_num
        jud_cc = case_instance.docketed_judgment_cc
        jud_yy = case_instance.docketed_judgment_yy
        docket_num = "{}-{}-{}{}".format(judgment_type_code, seq_num, jud_cc, jud_yy)
        return docket_num
        # return "{}-{}-{}{}".format(case_instance.judgment_type_code,
        #                            case_instance.seq_num,
        #                            case_instance.docketed_judgment_cc,
        #                            case_instance.docketed_judgment_yy)

    @staticmethod
    def bk_judgment_nums_from_cases(cases):
        """
        Return a list of bankruptcy numbers associated with a list o cases
        :param cases: a list of cases to use for processing
        :return: a list of bankruptcy numbers
        """
        bkcy_list = []
        for case_instance in cases:
            bkcy_case = case_instance.find('BANKRUPTCY_NUMBER')
            if not bkcy_case is None:
                bkcy_list.append(bkcy_case.text)
        return bkcy_list

    @staticmethod
    def usdc_judgment_nums_from_cases(cases):
        """
        Return a list of USDC numbers associated with a list o cases
        :param cases: a list of cases to use for processing
        :return: a list of USDC numbers
        """
        usdc_list = []
        for case_instance in cases:
            usdc_case = case_instance.find('USDC_CASE_NUMBER')
            if not usdc_case is None:
                usdc_num = case_instance.find('USDC_CASE_NUMBER').text
                usdc_list.append(usdc_num)
        return usdc_list

    @staticmethod
    def calc_xlsx_page_offset(xlsx_filename):
        """
        Figure out what the last page number is in an excel file
        :param xlsx_filename: the excel file to process
        :return: the last page number referenced in the excel file
        """
        import os
        from openpyxl import load_workbook, Workbook

        xlsx_pathname = os.path.join(settings.MEDIA_ROOT, xlsx_filename)
        last_page_num = 0
        PAGE_NUM_COL = 'K'  # page number column
        if not os.path.exists(xlsx_pathname):
            return last_page_num
        try:
            wb = load_workbook(xlsx_pathname)
            ws = wb.worksheets[0]
            for i in range(1, len(ws[PAGE_NUM_COL])):  # skip header
                current_row_page_num = int(ws[PAGE_NUM_COL][i].value)
                if current_row_page_num > last_page_num:
                    last_page_num = current_row_page_num
        except OSError:
            print("An error occurred while determining the page offset during BKCY XLSX operation")
            pass
        return last_page_num

    @staticmethod
    def check_reset_csv_file_data(order_id):
        # determine whether this search has been executed before or not
        # if not, pass
        # if so, was it for this search or a previous search that's being regenerated
        # if for a regenerated search, clear old file name (already being done)
        # so @TODO: append data for all instances except when filename already exists
        pass
